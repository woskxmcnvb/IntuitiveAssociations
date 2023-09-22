import io
import os

import pandas as pd 
import numpy as np
from datetime import datetime
import xlsxwriter

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

import matplotlib.pyplot as plt 
import seaborn as sns
    

class IADatabase: 
    df_ = None
    
    job_type_categories_ = ['ad.look', 'adhoq']
    ad_brand_categories_ = ['ad', 'brand']
    wtype_categories_ =    ['warm_up', 'active']
    
    time_from_ = 150
    time_to_ = 2500 
    
    COLUMNS_ = ['JOB_ID', 'JOB_TYPE', 'QST_NO', 
                'IA_CELL', 'IA_ORD', 'IA_WORDS', 'IA_MS', 'IA_ANSWER', 'IA_ATTEMP', 'IA_AD_BRAND', 'IA_WTYPE']
    
    
    def __init__(self, *args, **kwargs):
        self.df_ = pd.DataFrame({
            'JOB_ID':       pd.Series(dtype = 'int'),
            'JOB_TYPE':     pd.Categorical([], categories=self.job_type_categories_, ordered=False),
            'QST_NO':       pd.Series(dtype = 'int'),
            'IA_CELL':      pd.Series(dtype = 'int'),
            'IA_ORD':       pd.Series(dtype = 'int'),
            'IA_WORDS':     pd.Categorical([]),
            'IA_MS':        pd.Series(dtype = 'int'),
            'IA_ANSWER':    pd.Series(dtype = 'int'),
            #'IA_ANSWER':    pd.Categorical([], categories=['yes', 'no'], ordered=False),
            'IA_ATTEMP':    pd.Series(dtype = 'int'),
            'IA_AD_BRAND':  pd.Categorical([], categories=self.ad_brand_categories_, ordered=False),
            'IA_WTYPE':     pd.Categorical([], categories=self.wtype_categories_, ordered=False)
        })

    def __getattr__(self, attr):
        return getattr(self.df_, attr)
    
    def Deserialize(self, file_name): 
        self.df_ = pd.read_pickle(file_name)
        
    def Size(self): 
        return len(self.df_)
    
    def Serialize(self, file_name): 
        self.df_.to_pickle(file_name)
        
    def IsJobInDatabase(self, job_id): 
        if job_id in self.df_['JOB_ID'].unique(): 
            return True
        else: 
            return False
        
    def DropJobs(self, job_list): 
        self.df_ = self.df_[~self.df_['JOB_ID'].isin(job_list)]
        
    
    def AppendNewData(self, new_data):
        #################
        ## проверить что все столбцы на месте - нужно сделать 
        #################
        
        self.df_['IA_WORDS'] = self.df_['IA_WORDS'].cat.set_categories(
            set(new_data['IA_WORDS'].cat.categories) | set(self.df_['IA_WORDS'].cat.categories)
        )
        new_data['IA_WORDS'] = new_data['IA_WORDS'].cat.set_categories(
            set(new_data['IA_WORDS'].cat.categories) | set(self.df_['IA_WORDS'].cat.categories)
        )
        
        self.df_ = pd.concat([self.df_, new_data[self.COLUMNS_]])
        
    def GetShitFilter(self): 
        return (#self.df_['IA_ANSWER'].isin([1,2]) & 
                (self.df_['IA_WTYPE'] == 'active') & 
                (self.df_['IA_MS'] > self.time_from_) & 
                (self.df_['IA_MS'] < self.time_to_))
        
    def RespondentSpeed(self, job_type, ad_brand): 
        selection = ['JOB_ID', 'QST_NO', 'IA_MS']
        
        _filter = ( self.GetShitFilter() & 
                   (self.df_['JOB_TYPE'].isin(job_type)) & 
                   (self.df_['IA_AD_BRAND'].isin(ad_brand)))
            
        _by_respondent = self.df_.loc[_filter, selection].groupby(['JOB_ID', 'QST_NO']).mean()
        
        return _by_respondent.mean()[0] #, _by_respondent.std()[0]
    
    def GetNorms(self): 
        grouper = ['JOB_TYPE', 'IA_AD_BRAND', 'IA_ANSWER']
        selection = grouper + ['IA_MS']
        
        return self.loc[self.GetShitFilter(), selection].groupby(grouper).agg(['mean', 'std'])
    
    def JobList(self):
        return self.df_['JOB_ID'].unique()
    
    
    
    
    


class IAReporter: 
    database_ = IADatabase()
    report_file_ = 'ad_report.xlsx'
    temp_files_ = [] 
        
    def ReadDataFile(self, file_name, job_id, job_type):
        if job_id < 10000 or job_id > 99999:
            print("ОШИБКА ЧТЕНИЯ: job_id ожидается 5-значным")
            return None
        
        if job_type not in self.database_.job_type_categories_: 
            print("ОШИБКА ЧТЕНИЯ: job_type ожидается ['ad.look', 'adhoq']")
            return None
        
        new_data = pd.read_excel(
            file_name, 
            dtype={
                'QST_NO':       'int',
                'IA_CELL':      'int', 
                'IA_ORD':       'int',
                'IA_WORDS':     'category',
                'IA_MS':        'int',
                'IA_ANSWER':    'int',
                #'IA_ANSWER':    'category',
                'IA_ATTEMP':    'object',
                'IA_AD_BRAND':  'category',
                'IA_WTYPE':     'category'
            }
        )
        
        #################
        ## проверка типов нужно сделать 
        #################
        
        new_data['JOB_ID'] = job_id
        new_data['JOB_TYPE'] = pd.Categorical(
            len(new_data) * [job_type], categories=self.database_.job_type_categories_, 
            ordered=False) 
        
        for col in self.database_.columns: 
            if col not in new_data.columns: 
                print("ОШИБКА ЧТЕНИЯ: Не хватает столбца " + col)
                return None
        
        if not all(x in self.database_.ad_brand_categories_ for x in new_data['IA_AD_BRAND'].cat.categories): 
            print("ОШИБКА ЧТЕНИЯ: Неправильные категории в столбце IA_AD_BRAND")
            return None
        new_data['IA_AD_BRAND'] = new_data['IA_AD_BRAND'].cat.set_categories(self.database_.ad_brand_categories_, ordered=False)
        
        if not all(x in self.database_.wtype_categories_ for x in new_data['IA_WTYPE'].cat.categories): 
            print("ОШИБКА ЧТЕНИЯ: Неправильные категории в столбце IA_WTYPE")
            return None 
        
        new_data['IA_WTYPE'] = new_data['IA_WTYPE'].cat.set_categories(self.database_.wtype_categories_, ordered=False)
        new_data.loc[new_data['IA_ATTEMP'].isna(), 'IA_ATTEMP'] = 99
        new_data['IA_ATTEMP'] = new_data['IA_ATTEMP'].astype('int')
        
        return new_data
    
    def BuildJobReport(self, file_name, job_id, job_type): 
        ad = self.ReadDataFile(file_name, job_id, job_type)
        if ad is None: 
            print("Прервано")
            return
        
        self.database_.Deserialize('dump.pickle')
        
        if self.database_.IsJobInDatabase(job_id): 
            print("Этот проект уже в базе")
        else: 
            self.database_.AppendNewData(ad)
            print("Этого проекта нет в базе. Добавлено")
            self.database_.Serialize('dump.pickle')
        
        
        shit_filter = (
                (ad['IA_ANSWER'] != 3) & 
                (ad['IA_WTYPE'] != 'warm_up') & 
                (ad['IA_MS'] > self.database_.time_from_) & 
                (ad['IA_MS'] < self.database_.time_to_))
        
        ad['no_shit'] = shit_filter 
        
        # индивидуальная скорость респондента
        ad['resp_speed'] = ad.loc[shit_filter, ['QST_NO', 'IA_MS']].groupby(['QST_NO']).transform('mean')
        ad['resp_speed_norm'] = self.database_.RespondentSpeed(ad['JOB_TYPE'].unique(), ad['IA_AD_BRAND'].unique())
        ad['resp_speed_koef'] = ad['resp_speed']  / ad['resp_speed_norm'] 
        
        
        # подтягиваем нормы
        norms = self.database_.GetNorms() 
        
        ad = ad.merge(norms['IA_MS'].reset_index(), 
               how='left', 
               on=['JOB_TYPE', 'IA_AD_BRAND','IA_ANSWER'])
        
        # быстро или медленно
        ad['is_fast'] = pd.Categorical(['slow'] * len(ad), categories=['fast', 'slow'])
        ad.loc[
            ad['IA_MS'] < ad['mean'] * ad['resp_speed_koef'] - 0.5 * ad['std'] * (ad['resp_speed_koef'] ** 0.5), 
            'is_fast'] = 'fast'
        
        
        # Отчет в файл
        report_file = str(job_id) + self.report_file_
        
        if os.path.isfile(report_file):
            os.remove(report_file)
        
        # таблицы
        writer = pd.ExcelWriter(report_file, engine='xlsxwriter')
        norms.to_excel(writer, sheet_name='norms')
        ad.to_excel(writer, index=False, sheet_name='data')
        writer.close()
        
        # описательный отчет
        self.DescriptiveReport(job_id, ad)
        self.IACharts(job_id, ad)
        
        print('Отчет готов')
        return ad

    def SaveFigToSheet(self, fig, title, sheet, row=0, column='A'):  
        temp_file = 'temp' + str(len(self.temp_files_)) + '.png'
        self.temp_files_.append(temp_file)
        
        plt.tight_layout()
        fig.savefig(temp_file, format='png')
        
        sheet[column + str(row)] = title
        sheet.add_image(openpyxl.drawing.image.Image(temp_file), column + str(row + 1))
        plt.close(fig)
        
    def CleanTempFiles(self):
        for f in self.temp_files_:
            os.remove(f)
        self.temp_files_.clear()
        
            
    def IACharts(self, job_id, ad):
        report_file = str(job_id) + self.report_file_
        
        workbook = openpyxl.load_workbook(report_file)
        if 'key charts' in workbook.sheetnames:
            del workbook['key charts']
        key_sheet = workbook.create_sheet('key charts')
        
        
        ia_table = pd.concat(
            [
                pd.pivot_table(ad[ad['no_shit']], 
                       index=['IA_CELL', 'IA_AD_BRAND', 'IA_WORDS'], 
                       values='IA_ANSWER', 
                       aggfunc=lambda x: x.value_counts()[1] / len(x)),

                pd.pivot_table(ad[ad['no_shit'] & (ad['IA_ANSWER'] == 1)], 
                       index=['IA_CELL', 'IA_AD_BRAND', 'IA_WORDS'], 
                       values='is_fast', 
                       aggfunc=lambda x: x.value_counts()['fast'] / len(x))
            ], 
            axis=1
        )
        
        for r in dataframe_to_rows(ia_table, index=True, header=True):
            key_sheet.append(r)

        for cell in key_sheet['A'] + key_sheet[1]:
            cell.style = 'Pandas'
        
        
        begin_row = 1
        for cell in ad['IA_CELL'].unique(): 
            ia_table = pd.concat(
                [
                    pd.pivot_table(ad[ad['no_shit'] & (ad['IA_CELL'] == cell)], 
                           index=['IA_AD_BRAND', 'IA_WORDS'], 
                           values='IA_ANSWER', 
                           aggfunc=lambda x: x.value_counts()[1] / len(x)),

                    pd.pivot_table(ad[ad['no_shit'] & (ad['IA_CELL'] == cell) & (ad['IA_ANSWER'] == 1)], 
                           index=['IA_AD_BRAND', 'IA_WORDS'], 
                           values='is_fast', 
                           aggfunc=lambda x: x.value_counts()['fast'] / len(x))
                ], 
                axis=1
            ).reset_index().rename(
                columns={'IA_AD_BRAND': 'section', 'IA_WORDS': 'word', 'IA_ANSWER': 'percent yes', 'is_fast': 'percent fast'}
            )

            chart_table = ia_table[ia_table['section'] == 'brand']
        
            fig, ax = plt.subplots(1, 1, figsize=(10, 10))
            ax.scatter(chart_table['percent yes'], chart_table['percent fast'])
            ax.set_xlabel('% yes')
            ax.set_ylabel('% fast yes')
            for w, coord in chart_table[['word', 'percent yes', 'percent fast']].iterrows():
                ax.annotate(coord['word'], coord[['percent yes', 'percent fast']])

            self.SaveFigToSheet(fig, '', key_sheet, begin_row, 'M')
            begin_row += 40
        
        workbook.save(report_file)
        workbook.close() 
        self.CleanTempFiles()
        
            
    def DescriptiveReport(self, job_id, ad): 
        report_file = str(job_id) + self.report_file_
        
        workbook = openpyxl.load_workbook(report_file)
        if 'descriptive' in workbook.sheetnames:
            del workbook['descriptive']
        descriptive_sheet = workbook.create_sheet('descriptive')
        
        
        # Гистограммы времени ответа
        fig, (ax1, ax2, ax3, ax4) = plt.subplots(4, 1, figsize=(10, 10))
        sns.histplot(ad, x="IA_MS", ax=ax1, bins=100)
        sns.histplot(ad[ad["IA_MS"] < 3000], x="IA_MS", ax=ax2, bins=100)
        sns.histplot(ad[ad["IA_MS"] < 500], x="IA_MS", ax=ax3, bins=100)
        sns.histplot(ad[ad['no_shit']], x="IA_MS", ax=ax4, bins=100, label='Без хвостов', color='red')

        plt.legend()

        ax1.set_xlabel('')
        ax2.set_xlabel('')
        ax3.set_xlabel('')
        ax4.set_xlabel('')

        self.SaveFigToSheet(fig, 'Гистограммы времени ответа', descriptive_sheet, 1)

        #Зависимость от номера слова 
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 5))

        pd.pivot_table(ad, index=['IA_AD_BRAND', 'IA_ORD'], values='IA_MS', aggfunc=np.mean)['IA_MS'].\
            plot(ax=ax1, label='Время отклика')

        pd.pivot_table(ad, index=['IA_AD_BRAND', 'IA_ORD'], values='IA_ANSWER', 
                       aggfunc=lambda x: x.value_counts()[1] / len(x))['IA_ANSWER'].\
            plot(kind='bar', ax=ax2, label='% да')

        ax1.legend()
        ax2.legend()
        ax1.set_xlabel('')
        ax2.set_xlabel('')
        
        self.SaveFigToSheet(fig, 'Зависимость от номера слова', descriptive_sheet, 40)

        # Разброс по респондентам 
        fig, (ax1) = plt.subplots(1, 1, figsize=(10, 2))
        ad[['QST_NO', 'IA_MS']].groupby("QST_NO").mean()['IA_MS'].plot(kind='hist', ax=ax1, bins=50)
        
        self.SaveFigToSheet(fig, 'Разброс по респондентам', descriptive_sheet, 60)
        
        
        
        # Ответы да Нет  
        fig, axs = plt.subplots(2, 2, figsize=(10, 10))

        sns.countplot(data=ad, y="IA_ANSWER", ax=axs[0, 0])
        axs[0, 0].set_title('Количество ответов Да, Нет, Хз')
        axs[0, 0].set_xlabel('')
        axs[0, 0].set_ylabel('')

        sns.barplot(data=ad, x="IA_ANSWER", y="IA_MS", ax=axs[0, 1])
        axs[0, 1].set_title('Время ответов Да, Нет, Хз')
        axs[0, 1].set_xlabel('')
        axs[0, 1].set_ylabel('')

        sns.barplot(data=ad[ad["IA_ANSWER"] != 3], x="IA_ANSWER", y="IA_MS", ax=axs[1, 1])
        axs[1, 1].set_title('Время ответов Да, Нет')
        axs[1, 1].set_xlabel('')
        axs[1, 1].set_ylabel('')

        sns.histplot(data=ad[ad["IA_ANSWER"] == 1], x="IA_MS", ax=axs[1, 0])
        sns.histplot(data=ad[ad["IA_ANSWER"] == 2], x="IA_MS", ax=axs[1, 0], color='r')
        axs[1, 0].set_title('Распределение Да, Нет по времени')
        axs[1, 0].set_xlabel('')
        axs[1, 0].set_ylabel('')
        
        self.SaveFigToSheet(fig, 'Ответы Да, Нет', descriptive_sheet, 70)
        
        
        # По словам 
        #plt.tight_layout()

        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 10))
        ad[['IA_MS', 'IA_WTYPE', 'IA_WORDS', 'IA_AD_BRAND']].\
            groupby(['IA_AD_BRAND', 'IA_WTYPE', 'IA_WORDS'], observed=True).mean().plot.bar(y="IA_MS", ax=ax2)


        ad[['IA_MS', 'IA_AD_BRAND', 'IA_WTYPE']].\
            groupby(['IA_AD_BRAND', 'IA_WTYPE'], observed=True).mean().plot.bar(y="IA_MS", ax=ax1)
        
        ax1.set_xlabel('')
        ax2.set_xlabel('')
        
        self.SaveFigToSheet(fig, 'По словам', descriptive_sheet, 110)
        
        workbook.save(report_file)
        workbook.close() 
        self.CleanTempFiles()
            
            
    