from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

import pandas as pd

class ExcelReportBuilder:


    def __init__(self, file_name): 
        self.images_ = []
        self.tables_ = []
        self.file_name_ = file_name 

    def AddImage(self, stream, page_name: str, position: str="A1"):
        self.images_.append({'page': page_name, 'image': Image(stream), 'position': position})
    
    def AddTable(self, table: pd.DataFrame, page_name: str, drop_index=False):
        self.tables_.append({'page': page_name, 'table': table, 'index': not(drop_index)})

    @staticmethod
    def __GetSheetPtr(wb: Workbook, name: str):
        if name not in wb.sheetnames:
            return wb.create_sheet(name)
        else: 
            return wb[name]

    
    def SaveToFile(self):
        wb = Workbook()
        
        for tab in self.tables_:
            ws = ExcelReportBuilder.__GetSheetPtr(wb, tab['page'])
            for r in dataframe_to_rows(tab['table'], index=tab['index'], header=True):
                ws.append(r)

        for img in self.images_:
            ws = ExcelReportBuilder.__GetSheetPtr(wb, img['page'])
            ws.add_image(img['image'], img['position'])

        del wb['Sheet']
        wb.save(self.file_name_)

